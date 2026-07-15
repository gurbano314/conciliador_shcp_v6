# ============================================================
# CONCILIADOR DE COTIZACIONES  |  engine.py  v6.4
# Motor de extracción puro – sin dependencias de GUI.
# Compatible con PyQt6 y cualquier otro frontend.
# ============================================================
from __future__ import annotations

import datetime
import hashlib
import io
import math
import re
import sys
from typing import Optional

import fitz          # PyMuPDF
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# CONSTANTES Y REGEX
# ─────────────────────────────────────────────────────────────
NATIVE_MIN_CHARS_PER_PAGE = 80
MAX_PLAUSIBLE_MXN = 500_000_000  # 500 M – presupuestos de obra pueden superar 50 M

_BANXICO_SERIES = {"USD": "SF43718", "EUR": "SF46410", "CAD": "SF60653"}
_BANXICO_URL = "https://www.banxico.org.mx/SieAPIRest/service/v1"

_COLS = [
    "Nº Sec", "Sección", "Tipo",
    "Fecha", "Rubro", "QT", "T. Cambio", "(+ IVA)", "Cantidad",
    "Precio Unitario", "Subtotal (Sin IVA)", "IVA 16%", "Total con IVA",
    "Diferencia final", "Monto en Anexo Escrito", "Observaciones",
]
_WIDTHS_COL = [8, 12, 20, 12, 52, 5, 10, 7, 8, 16, 18, 17, 16, 18, 22, 48]

_MONEY_RE = re.compile(
    r"MX\$\s*([\d,]+(?:\.\d{1,2})?)"              # g1: MX$ prefix
    r"|\$\s*([\d,]+(?:\.\d{1,2})?)"               # g2: $ prefix (acepta $15,000 y $15000)
    r"|(?<!\d)([\d]{1,3}(?:,\d{3})+(?:\.\d{1,2})?)"  # g3: número con comas (15,000)
    r"|(?<!\d)(\d{1,9}\.\d{2})(?!\d)"             # g4: decimal suelto (15000.00)
    r"|(?:(?:^|(?<=[:\s]))\s*([1-9]\d{3,8})(?=\s*$|\s+[^\d]))"  # g5: entero grande sin coma (15000)
)
_MONETARY_CTX = re.compile(
    r"total|importe|monto|precio|valor|costo|cobro|cargo|pago"
    r"|subtotal|honorarios|tarifa|renta|fianza",
    re.IGNORECASE,
)
_DATE_RE = re.compile(
    r"(\d{1,2})\s*(?:de)?\s*([a-záéíóúñA-ZÁÉÍÓÚÑ]+)[,\s]*(?:del?\s*)?(\d{4})"
    r"|(\d{4})[\s.\-/]+(\d{1,2})[\s.\-/]+(\d{1,2})"
    r"|(\d{1,2})[\s.\-/]+(\d{1,2})[\s.\-/]+(\d{2,4})",
    re.IGNORECASE,
)
_ITEM_RE = re.compile(
    r"^\d+\s+(\d+)\s+\w+\s+(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$"
)
_GHOST_MONEY_RE = re.compile(
    r"(?:^|\s)\$?\s*([\d]{1,3}(?:,[\d]{3})*(?:\.\d{1,2}))"
    r"|(?:^|\s)([\d]{1,9}\.\d{2})(?!\d)",
    re.IGNORECASE,
)
_IVA_INCLUDED_RE = re.compile(
    r"iva\s+inclu[iy]|inclu[iy].*iva|ya\s+inclu|tarifa\s+inclu|con\s+iva\s+inclu|precio\s+(?:total|final)\s+con\s+iva",
    re.IGNORECASE,
)
_IVA_EXEMPT_RE = re.compile(
    r"exento|tasa\s+0|no\s+aplica\s+iva|sin\s+iva|iva\s*:\s*0|libre\s+de\s+iva"
    r"|iva\s+no\s+aplica|no\s+causa\s+iva",
    re.IGNORECASE,
)
_ES_DE_RE = re.compile(
    r"(?:es|ser[aá]|era|monto)\s+de\s+\$?\s*([\d,]+(?:\.\d{1,2})?)",
    re.IGNORECASE,
)
_TBL_TOTAL_RE  = re.compile(r"\btotal\b", re.IGNORECASE)
_TBL_SUBTOT_RE = re.compile(r"sub\s*-?\s*total|sin\s*iva|importe\s*neto", re.IGNORECASE)
_TBL_IVA_RE    = re.compile(r"\biva\b|i\.?\s*v\.?\s*a\.?|16\s*%|vat", re.IGNORECASE)
_TBL_EXCL_RE   = re.compile(r"sub|parcial|acum", re.IGNORECASE)
_BUDGET_EXCL   = re.compile(
    r"presupuestal|presupuesto\s+total"
    r"|costo\s+(?:total\s+)?del?\s+proyecto"
    r"|presupuesto\s+del?\s+proyecto"
    r"|inversi[óo]n\s+total\s+del\s+(?:proyecto|programa)"  # NO excluir "del servicio"
    r"|anexo\s*\d|aportante",
    re.IGNORECASE,
)
# Líneas de plan de pagos – no confundir con el total global
_PAYMENT_SCHED_RE = re.compile(
    r"pago\s+mensual|mensualidad|por\s+mes\s*\("
    r"|cantidad\s+de\s+pagos|n[uú]mero\s+de\s+pagos"
    r"|cuota\s+mensual|abono\s+mensual",
    re.IGNORECASE,
)

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}

_FMT_MONEY = (
    '_-"$ "* #,##0.00_-;'
    '\\-"$ "* #,##0.00_-;'
    '_-"$ "* "-"??_-;_-@_-'
)
_FMT_DATE = "mm-dd-yy"

# Caché Banxico a nivel de módulo (compartido entre sesiones del mismo proceso)
_BX_CACHE: dict[tuple, Optional[float]] = {}

# OCR – singleton a nivel de módulo para evitar recargar el modelo
_ocr_engine = None
_ocr_available: Optional[bool] = None   # None = no determinado aún


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def _md5(b: bytes) -> str:
    return hashlib.md5(b).hexdigest()


def _is_blank_value(v) -> bool:
    if v is None:
        return True
    try:
        if bool(pd.isna(v)):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(v, str) and not v.strip()


def _normalize_editor_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame(columns=_COLS)
    clean = pd.DataFrame(df).copy()
    private_cols = [c for c in clean.columns if str(c).startswith("_")]
    clean = clean.drop(columns=private_cols, errors="ignore").reindex(columns=_COLS)
    if clean.empty:
        return clean.reset_index(drop=True)
    editable_cols = [c for c in _COLS if c != "Diferencia final"]
    blank_rows = clean[editable_cols].apply(
        lambda row: all(_is_blank_value(v) for v in row), axis=1
    )
    return clean.loc[~blank_rows].reset_index(drop=True)


def _df_hash(df: pd.DataFrame) -> str:
    if df is None:
        return ""
    df = _normalize_editor_df(df)
    hash_df = df.astype(object).where(pd.notna(df), "")
    try:
        return hashlib.md5(
            pd.util.hash_pandas_object(hash_df, index=False).values.tobytes()
        ).hexdigest()
    except Exception:
        return hashlib.md5(hash_df.to_csv(index=False).encode()).hexdigest()


def _safe_f(v) -> Optional[float]:
    if v is None:
        return None
    try:
        f = float(str(v).replace(",", "").strip())
        return None if math.isnan(f) or math.isinf(f) else f
    except (ValueError, TypeError):
        return None


def _money(txt: str) -> Optional[float]:
    """Devuelve el primer monto positivo encontrado en txt."""
    for m in _MONEY_RE.finditer(str(txt)):
        raw = m.group(1) or m.group(2) or m.group(3) or m.group(4) or m.group(5)
        if raw:
            try:
                v = float(raw.replace(",", ""))
                if v > 0:
                    return v
            except ValueError:
                continue
    return None


def _money_max(txt: str) -> Optional[float]:
    """Devuelve el MAYOR monto positivo encontrado en txt (útil para líneas multivalor)."""
    best: Optional[float] = None
    for m in _MONEY_RE.finditer(str(txt)):
        raw = m.group(1) or m.group(2) or m.group(3) or m.group(4) or m.group(5)
        if raw:
            try:
                v = float(raw.replace(",", ""))
                if v > 0 and (best is None or v > best):
                    best = v
            except ValueError:
                continue
    return best


def _date(text: str) -> Optional[datetime.date]:
    for m in _DATE_RE.finditer(text.lower()):
        g = m.groups()
        try:
            if g[0]:
                mes_str = g[1].lower().strip()
                mes_num = _MESES.get(mes_str) or _MESES.get(mes_str[:3])
                if not mes_num:
                    for k, v in _MESES.items():
                        if mes_str.startswith(k[:3]):
                            mes_num = v
                            break
                if mes_num:
                    return datetime.date(int(g[2]), mes_num, int(g[0]))
            if g[3]:
                return datetime.date(int(g[3]), int(g[4]), int(g[5]))
            if g[6]:
                yr = int(g[8])
                return datetime.date(yr + 2000 if yr < 100 else yr, int(g[7]), int(g[6]))
        except (ValueError, KeyError):
            continue
    return None


# ─────────────────────────────────────────────────────────────
# CONSOLIDACIÓN DE PDFs
# ─────────────────────────────────────────────────────────────
def build_combined_pdf(file_list: list[dict]) -> tuple[bytes, str, int, list[dict]]:
    """Combina varios PDFs en uno solo y devuelve (bytes, hash, total_pages, page_map)."""
    combined = fitz.open()
    page_map: list[dict] = []
    for file_idx, f_info in enumerate(file_list):
        src = fitz.open(stream=f_info["bytes"], filetype="pdf")
        n = len(src)
        combined.insert_pdf(src)
        for local_p in range(n):
            page_map.append({
                "file_idx":   file_idx,
                "local_page": local_p,
                "name":       f_info["name"],
            })
        src.close()
    buf = io.BytesIO()
    combined.save(buf)
    combined.close()
    combined_bytes = buf.getvalue()
    return combined_bytes, _md5(combined_bytes), len(page_map), page_map


# ─────────────────────────────────────────────────────────────
# RENDER DE PÁGINA (sin caché – la gestiona el caller)
# ─────────────────────────────────────────────────────────────
def render_page(pdf_bytes: bytes, idx: int, scale: float = 1.5) -> bytes:
    """Renderiza la página idx del PDF a PNG. Devuelve b'' en caso de error."""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            if idx < 0 or idx >= len(doc):
                return b""
            pix = doc[idx].get_pixmap(
                matrix=fitz.Matrix(scale, scale), colorspace=fitz.csRGB, alpha=False
            )
            return pix.tobytes("png")
    except Exception:
        return b""


# ─────────────────────────────────────────────────────────────
# OCR
# ─────────────────────────────────────────────────────────────
def _get_ocr():
    """Devuelve el motor OCR (singleton). None si no está disponible.

    Soporta dos generaciones de rapidocr-onnxruntime:
    - v1.3.x+ (API nueva): RapidOCR(det_model_dir=..., rec_model_dir=..., cls_model_dir=...)
    - v1.2.x  (API vieja, Python 3.14): RapidOCR() sin argumentos de directorio
    """
    global _ocr_engine, _ocr_available
    if _ocr_available is False:
        return None
    if _ocr_engine is not None:
        return _ocr_engine
    try:
        from rapidocr_onnxruntime import RapidOCR
        try:
            # API 1.2.x o la más nueva por defecto (sin parámetros)
            _ocr_engine = RapidOCR()
        except Exception:
            # Fallback para algunas versiones intermedias (1.3.x) que exigían los argumentos
            _ocr_engine = RapidOCR(
                det_model_dir=None, rec_model_dir=None, cls_model_dir=None
            )
        _ocr_available = True
        return _ocr_engine
    except Exception as exc:
        print(f"[engine] RapidOCR no disponible: {exc}", file=sys.stderr)
        _ocr_available = False
        return None


def ocr_page(
    pdf_bytes: bytes,
    idx: int,
    on_warning: Optional[callable] = None,
) -> str:
    """OCR de una página. on_warning(msg) se llama ante errores no fatales.

    Renderiza a 2.5x para mejorar la lectura de documentos escaneados
    con baja resolución (ej. fotos de celular o escaneos oscuros).
    """
    ocr = _get_ocr()
    if ocr is None:
        return ""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            pix = doc[idx].get_pixmap(
                matrix=fitz.Matrix(2.0, 2.0), colorspace=fitz.csRGB, alpha=False
            )
            img = np.frombuffer(pix.samples, np.uint8).reshape(
                pix.height, pix.width, 3
            )[:, :, ::-1]
            result, _ = ocr(img)
            if result:
                return "\n".join(r[1] for r in result if r and len(r) > 1)
            return ""
    except Exception as exc:
        msg = f"Error OCR pág. {idx + 1}: {exc}"
        if on_warning:
            on_warning(msg)
        else:
            print(f"[engine] {msg}", file=sys.stderr)
        return ""


# ─────────────────────────────────────────────────────────────
# API BANXICO
# ─────────────────────────────────────────────────────────────
def banxico_tc(
    moneda: str,
    fecha: datetime.date,
    token: str,
    on_warning: Optional[callable] = None,
) -> Optional[float]:
    """Tipo de cambio Banxico. Usa caché de módulo. Emite on_warning ante errores."""
    if moneda not in _BANXICO_SERIES or not token:
        return None
    cache_key = (moneda, fecha.isoformat())
    if cache_key in _BX_CACHE:
        return _BX_CACHE[cache_key]
    serie = _BANXICO_SERIES[moneda]
    f_ini = (fecha - datetime.timedelta(days=5)).isoformat()
    url = f"{_BANXICO_URL}/series/{serie}/datos/{f_ini}/{fecha.isoformat()}"
    try:
        resp = requests.get(
            url,
            headers={"Bmx-Token": token, "Accept": "application/json"},
            timeout=10,
        )
        resp.raise_for_status()
        datos = resp.json()["bmx"]["series"][0]["datos"]
        if not datos:
            _BX_CACHE[cache_key] = None
            return None
        tc = float(datos[-1]["dato"].replace(",", ""))
        _BX_CACHE[cache_key] = tc
        return tc
    except Exception as exc:
        msg = f"Banxico ({moneda} · {fecha}): {exc}"
        if on_warning:
            on_warning(msg)
        else:
            print(f"[engine] {msg}", file=sys.stderr)
        _BX_CACHE[cache_key] = None
        return None


# ─────────────────────────────────────────────────────────────
# RECALCULAR COLUMNAS DERIVADAS
# ─────────────────────────────────────────────────────────────
def recalc_derived(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recalcula columnas derivadas conservadoramente:
    solo rellena campos vacíos (None/NaN), nunca sobreescribe
    valores que el usuario ya proporcionó o editó manualmente.
    """
    df = _normalize_editor_df(df)
    for idx, row in df.iterrows():
        qty = _safe_f(row.get("Cantidad"))
        qty = qty if qty and qty > 0 else 1

        unit = _safe_f(row.get("Precio Unitario"))
        sub  = _safe_f(row.get("Subtotal (Sin IVA)"))
        iva  = _safe_f(row.get("IVA 16%"))
        tot  = _safe_f(row.get("Total con IVA"))
        anx  = _safe_f(row.get("Monto en Anexo Escrito"))

        iva_flag = (
            str(row.get("(+ IVA)", "N/M") or "N/M").strip()
            .replace("\u00ed", "i").replace("\u00c3\u00ad", "i")
        ).lower()
        has_iva    = iva_flag in {"si", "yes", "true", "1", "sí"}
        iva_incl   = iva_flag == "incluido"
        iva_exento = iva_flag in {"exento", "no", "n/m", "nm", "n.a.", "na", ""}

        if sub is None and tot is not None and iva is not None:
            sub = round(tot - iva, 2)
        if unit is None:
            if sub is not None and qty > 0:
                unit = round(sub / qty, 2)
            elif tot is not None and qty > 0:
                base = (tot / 1.16) if (has_iva or iva_incl) else tot
                unit = round(base / qty, 2)
        if sub is None and unit is not None:
            sub = round(qty * unit, 2)
        if sub is not None:
            if iva is None:
                if has_iva:
                    iva = round(sub * 0.16, 2)
                elif iva_incl:
                    iva = round(sub - sub / 1.16, 2)
                elif iva_exento:
                    iva = 0.0
            if tot is None:
                if iva_incl and iva is not None:
                    tot = round(sub / 1.16 + iva, 2)
                else:
                    tot = round(sub + (iva or 0), 2)

        df.at[idx, "Cantidad"]           = int(qty) if float(qty).is_integer() else qty
        df.at[idx, "Precio Unitario"]    = unit
        df.at[idx, "Subtotal (Sin IVA)"] = sub
        df.at[idx, "IVA 16%"]            = iva
        df.at[idx, "Total con IVA"]      = tot
        df.at[idx, "Diferencia final"]   = (
            round((tot or 0) - (anx or 0), 2)
            if tot is not None or anx is not None else None
        )
    return df


# ─────────────────────────────────────────────────────────────
# GHOST GRID PARSER
# ─────────────────────────────────────────────────────────────
def _parse_space_table(text: str) -> list[dict]:
    items = []
    MIN_DESC = 4
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or _BUDGET_EXCL.search(stripped):
            continue
        m = _ITEM_RE.match(stripped)
        if m:
            qty_s, desc, pu_s, total_s = m.groups()
            items.append({
                "qty": int(qty_s), "desc": desc.strip(),
                "pu":  float(pu_s.replace(",", "")),
                "total": float(total_s.replace(",", "")),
            })
            continue
        montos = []
        for m2 in _GHOST_MONEY_RE.finditer(stripped):
            raw = m2.group(1) or m2.group(2)
            if raw:
                v = _safe_f(raw.replace(",", ""))
                if v and v >= 1.0:
                    montos.append(v)
        if not montos:
            continue
        first_match = _GHOST_MONEY_RE.search(stripped)
        desc_part = stripped[:first_match.start()].strip().rstrip(":-")
        if len(desc_part) < MIN_DESC:
            continue
        qty_d = 1
        desc_clean = desc_part
        qty_m = re.match(r"^(\d{1,4})\s+(.+)", desc_part)
        if qty_m and int(qty_m.group(1)) <= 9999:
            qty_d = int(qty_m.group(1))
            desc_clean = qty_m.group(2).strip()
        if len(montos) >= 2:
            montos_sorted = sorted(montos)
            pu_v  = montos_sorted[0]
            tot_v = montos_sorted[-1]
            tol = max(0.5, tot_v * 0.05)
            if abs(qty_d * pu_v - tot_v) <= tol or len(montos) == 2:
                items.append({"qty": qty_d, "desc": desc_clean, "pu": pu_v, "total": tot_v})
        elif len(montos) == 1:
            items.append({"qty": qty_d, "desc": desc_clean,
                          "pu": montos[0] / qty_d, "total": montos[0]})
    return items


# ─────────────────────────────────────────────────────────────
# MOTOR DE EXTRACCIÓN PRINCIPAL
# ─────────────────────────────────────────────────────────────
def extract(
    pdf_bytes: bytes,
    label: str,
    p0: int,
    p1: int,
    det_iva: bool,
    calc_sub: bool = True,
    tipo: str = "Cotización Proveedor",
    moneda: str = "AUTO",
    bx_token: str = "",
    sec_num: int = 1,
    on_progress: Optional[callable] = None,
    on_warning:  Optional[callable] = None,
) -> dict:
    """
    Motor de extracción de montos (5 estrategias + cota proporcional).
    on_progress(msg: str) – llamado al avanzar fases.
    on_warning(msg: str)  – llamado ante condiciones no fatales.
    """
    native_text = ""
    table_rows: list[list[str]] = []
    ocr_used = False

    # ── Estrategias 1 y 2: texto nativo + tablas ──────────────
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        n_pages = len(pdf.pages)
        pr = range(max(0, p0 - 1), min(p1, n_pages))
        for i in pr:
            pg = pdf.pages[i]
            txt = pg.extract_text() or ""
            native_text += "\n" + txt
            for tbl in pg.extract_tables() or []:
                if tbl:
                    table_rows.extend(
                        [str(c or "").replace("\n", " ").strip() for c in row]
                        for row in tbl if row
                    )
            for item in _parse_space_table(txt):
                table_rows.append([
                    str(item["qty"]), item["desc"],
                    f"${item['pu']:,.2f}", f"${item['total']:,.2f}",
                ])

    pages_count = max(len(pr), 1)
    if len(native_text.strip()) < NATIVE_MIN_CHARS_PER_PAGE * pages_count:
        ocr_used = True
        if on_progress:
            on_progress("OCR de páginas…")
        for i in pr:
            o_txt = ocr_page(pdf_bytes, i, on_warning=on_warning)
            if o_txt:
                native_text += "\n" + o_txt
                for item in _parse_space_table(o_txt):
                    table_rows.append([
                        str(item["qty"]), item["desc"],
                        f"${item['pu']:,.2f}", f"${item['total']:,.2f}",
                    ])

    text = native_text
    tlow = text.lower()

    # ── G1.5: Detectar Moneda ─────────────────────────────────
    if moneda == "AUTO":
        found_currency = True
        if re.search(r"\b(usd|dl[ls]s?|dollars?|dolares|dólares)\b", tlow):
            moneda = "USD"
        elif re.search(r"\b(eur|euros?|€)\b", tlow):
            moneda = "EUR"
        elif re.search(r"\b(cad|canadian)\b", tlow):
            moneda = "CAD"
        elif re.search(r"\b(mxn|pesos?|m\.n\.|mn)\b", tlow):
            moneda = "MXN"
        else:
            moneda = "MXN"
            found_currency = False
            
        if not found_currency and on_warning:
            on_warning(f"Sección '{label}': Moneda no explícita en texto. Se asume MXN. Por favor corrobore.")

    # ── G2: Detectar tipo de IVA ──────────────────────────────
    iva_mention = re.search(r"\biva\b|i\.?\s*v\.?\s*a\.?|16\s*%|vat|impuesto|exento|tasa\s*0", tlow)
    if not det_iva or not iva_mention:
        iva_f = "N/M"
    elif _IVA_EXEMPT_RE.search(text):
        iva_f = "Exento"
    elif _IVA_INCLUDED_RE.search(text):
        iva_f = "Incluido"
    else:
        iva_f = "Sí"

    tot = iva = sub = pu = fecha = None
    qty = 1
    obs_parts: list[str] = []
    if ocr_used:
        obs_parts.append("OCR")
    if iva_f in ("Exento", "Incluido"):
        obs_parts.append(f"IVA {iva_f}")

    # ── Estrategia 1: tablas estructuradas ───────────────────
    def _scan_table_rows(rows, tot_in, iva_in, sub_in):
        """
        Escanea filas de tabla buscando total / IVA / subtotal.
        Retorna (candidates, tot, iva, sub, sub_count).
        sub_count > 1 indica presupuesto multi-sección → omitir cota proporcional.
        """
        candidates: list[float] = []
        sub_count = 0
        for row in rows:
            j = "   ".join(row).lower()
            if _BUDGET_EXCL.search(j):
                continue
            if _TBL_TOTAL_RE.search(j) and not _TBL_EXCL_RE.search(j):
                # FIX: max por celda – tablas PRECIO|IVA|TOTAL tienen el mayor al final
                cell_moneys = [_money(c) for c in row if c]
                cell_moneys = [v for v in cell_moneys if v]
                v = max(cell_moneys) if cell_moneys else _money("   ".join(row))
                if v:
                    candidates.append(v)
                    if tot_in is None or v > tot_in:
                        tot_in = v
            if _TBL_IVA_RE.search(j):
                v = _money("   ".join(row))
                if v and (iva_in is None or v > iva_in):
                    iva_in = v
            if _TBL_SUBTOT_RE.search(j):
                v = _money("   ".join(row))
                if v:
                    if sub_in is None or v > sub_in:
                        sub_in = v
                    sub_count += 1   # cada subtotal de sección cuenta
        return candidates, tot_in, iva_in, sub_in, sub_count

    tot_candidates, tot, iva, sub, sub_count = _scan_table_rows(table_rows, tot, iva, sub)

    if sub is not None and len(tot_candidates) > 1:
        factor = 1.0 if iva_f in ("Exento", "N/M") else 1.16
        target = sub * factor
        best = min(tot_candidates, key=lambda x: abs(x - target))
        if abs(best - target) / max(target, 1) <= 0.20:
            tot = best
            obs_parts.append("Total seleccionado por coherencia")

    # ── Estrategia 2 (R1): triangulación qty × pu ≈ total ────
    valid_line_totals: list[float] = []
    seen_lines: set[tuple] = set()
    for row in table_rows:
        row_str = " ".join(str(c) for c in row)
        nums: list[float] = []
        for token in re.findall(r"[\d,]+(?:\.\d+)?", row_str):
            n = _safe_f(token.replace(",", ""))
            if n is not None and n > 0:
                nums.append(n)
        if len(nums) < 2:
            continue
        found = False
        for i_t, t_cand in enumerate(nums):
            if t_cand < 1.0 or found:
                continue
            for i_p, p_cand in enumerate(nums):
                if i_p == i_t or p_cand < 10.0 or found:
                    continue
                # Evitar p == t (ej: año "2026" duplicado = 1×2026≈2026)
                if abs(p_cand - t_cand) / max(t_cand, 1.0) < 0.001:
                    continue
                for i_q, q_cand in enumerate(nums):
                    if i_q in (i_t, i_p):
                        continue
                    if not (1 <= q_cand <= 9999 and q_cand == int(q_cand)):
                        continue
                    tol = max(0.5, t_cand * 0.01)
                    if abs(q_cand * p_cand - t_cand) <= tol:
                        key = (int(q_cand), round(p_cand, 2), round(t_cand, 2))
                        if key not in seen_lines:
                            seen_lines.add(key)
                            valid_line_totals.append(t_cand)
                        found = True
                        break
                if found:
                    break

    if tot is None and valid_line_totals:
        tot = round(sum(valid_line_totals), 2)
        obs_parts.append("Total por líneas")

    # ── G5: OCR selectivo para PDFs mixtos ────────────────────
    if not ocr_used and not table_rows and tot is None and sub is None:
        ocr_used = True
        if on_progress:
            on_progress("OCR selectivo (tabla imagen)…")
        for i in pr:
            o_txt = ocr_page(pdf_bytes, i, on_warning=on_warning)
            if o_txt:
                native_text += "\n" + o_txt
                for item in _parse_space_table(o_txt):
                    table_rows.append([
                        str(item["qty"]), item["desc"],
                        f"${item['pu']:,.2f}", f"${item['total']:,.2f}",
                    ])
        if ocr_used:
            text  = native_text
            tlow  = text.lower()
            obs_parts.append("OCR-Selectivo (tabla imagen)")
            if table_rows:
                _, tot, iva, sub, _sc2 = _scan_table_rows(table_rows, tot, iva, sub)
                sub_count += _sc2

    # ── Estrategia 3: texto libre ─────────────────────────────
    for ln in text.splitlines():
        if fecha is None:
            d = _date(ln)
            if d:
                fecha = d

    if tot is None:
        for ln in text.splitlines():
            if _BUDGET_EXCL.search(ln):
                continue
            if _PAYMENT_SCHED_RE.search(ln):    # ignorar "PAGO MENSUAL", etc.
                continue
            if re.search(r"\btotal\b", ln, re.I) and not re.search(r"sub|parcial|acum", ln, re.I):
                # FIX: máximo de la línea – cubre layout SUBTOTAL | IVA | TOTAL en columnas
                v = _money_max(ln)
                if v and (tot is None or v > tot):
                    tot = v

    # ── G3: Pares label:valor en líneas adyacentes ─────────────
    if tot is None or sub is None:
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        for i_ln, ln_a in enumerate(lines[:-1]):
            ln_b = lines[i_ln + 1]
            if _BUDGET_EXCL.search(ln_a):
                continue
            if _money(ln_a) is not None:
                continue
            v = _money(ln_b)
            if v is None:
                continue
            ln_a_low = ln_a.lower()
            # G3 puede sobreescribir tot si encuentra un valor etiquetado MÁS GRANDE
            # (cubre caso donde estrategia 2 detectó un total pequeño erróneo)
            if re.search(r"\btotal\b", ln_a_low) and \
               not re.search(r"sub|parcial|acum", ln_a_low) and \
               (tot is None or v > tot):
                tot = v
                obs_parts.append("Total (línea adyacente)")
            elif sub is None and re.search(
                r"sub\s*-?\s*total"
                r"|importe\s*(?:total|neto|final)|importe\s*s[ui]n\s*iva"
                r"|sin\s*iva",
                ln_a_low,
            ):
                sub = v
                obs_parts.append("Subtotal (línea adyacente)")

    # ── Estrategia 3.5: "es de $X" ───────────────────────────
    if tot is None:
        for ln in text.splitlines():
            m_ed = _ES_DE_RE.search(ln)
            if m_ed:
                v = _safe_f(m_ed.group(1).replace(",", ""))
                if v and 10.0 <= v <= MAX_PLAUSIBLE_MXN:
                    tot = v
                    obs_parts.append("Precio directo")
                    break

    # ── V1+O3: Cota de Cordura Proporcional ──────────────────
    if tot is not None:
        _iva_coherent = False
        if sub is not None and iva is not None and tot > 0:
            if abs(sub + iva - tot) / tot < 0.02:
                _iva_coherent = True
        # sub_count >= 2 → presupuesto multi-sección (SUBTOTAL A1, A2…)
        # En ese caso los subtotales son de sección, no de la cotización completa
        _multi_section = sub_count >= 2
        if not _iva_coherent and not _multi_section:
            if sub is not None and tot > sub * 1.5:
                obs_parts.append("⚠ Total descartado (desproporcionado vs subtotal)")
                tot = None
            # Nota: la cota por valid_line_totals fue eliminada por ser demasiado
            # agresiva en cotizaciones de multi-evento / multi-clínica.

    # ── Estrategia 4 (R2): fallback – máximo en contexto monetario
    if tot is None:
        all_vals: list[float] = []
        for ln in text.splitlines():
            if _BUDGET_EXCL.search(ln):
                continue
            if not _MONETARY_CTX.search(ln):
                continue
            for m in _MONEY_RE.finditer(ln):
                raw = m.group(1) or m.group(2) or m.group(3) or m.group(4)
                if raw:
                    try:
                        v = float(raw.replace(",", ""))
                        if 10.0 <= v <= MAX_PLAUSIBLE_MXN:
                            all_vals.append(v)
                    except ValueError:
                        pass
        if all_vals:
            if sub is not None:
                all_vals = [v for v in all_vals if v <= sub * 1.5]
            elif valid_line_totals:
                sum_lines = sum(valid_line_totals)
                all_vals = [v for v in all_vals if v <= sum_lines * 1.5]
            if all_vals:
                tot = max(all_vals)
                obs_parts.append("Total inferido (máx. validado)")

    if tot is not None and tot > MAX_PLAUSIBLE_MXN:
        # Solo avisar — no descartar. Presupuestos de obra pueden superar 500 M.
        obs_parts.append(f"⚠ Valor muy alto ({tot:,.2f}) — verificar")

    # Subtotal e IVA por texto libre
    if sub is None:
        for ln in text.splitlines():
            if re.search(r"sub\s*-?\s*total|importe|sin\s*iva", ln, re.I) and \
               not re.search(r"(?<!sub\s)(?<!sub)\btotal\b", ln, re.I):
                v = _money(ln)
                if v and (tot is None or v <= tot):
                    sub = v
                    break

    if iva is None and iva_f == "Sí":
        for ln in text.splitlines():
            if re.search(r"\biva\b|i\.?\s*v\.?\s*a\.?|16\s*%|vat|impuesto", ln, re.I):
                v = _money(ln)
                if v and (tot is None or v < tot):
                    iva = v
                    break

    # Cantidad y P.U.
    best_row_total = 0.0
    for row in table_rows:
        row_str = " ".join(str(c) for c in row)
        nums_: list[float] = []
        for t in re.findall(r"[\d,]+(?:\.\d+)?", row_str):
            n_val = _safe_f(t)
            if n_val is not None:
                nums_.append(n_val)
        # FIX: requiere ≥3 números para evitar que un dígito en la descripción
        # (ej: "5" en "TOTAL POR 5 CLÍNICAS") sea tomado como cantidad.
        if len(nums_) >= 3 and 1 <= nums_[0] <= 9999:
            row_tot = nums_[-1]
            if row_tot > best_row_total:
                best_row_total = row_tot
                qty = int(nums_[0])
                pu  = nums_[-2] if len(nums_) > 2 else nums_[-1]

    # ── Triangulación final (G2: 4 estados de IVA) ───────────
    if iva_f == "Incluido":
        if tot and not sub:
            sub = round(tot / 1.16, 2)
            iva = round(tot - sub, 2)
            obs_parts.append("IVA descompuesto del total")
        elif sub and not tot:
            tot = sub
            sub = round(tot / 1.16, 2)
            iva = round(tot - sub, 2)
    elif iva_f == "Exento":
        if tot and not sub:
            sub = tot
        elif sub and not tot:
            tot = sub
        iva = 0.0
    elif iva_f == "Sí":
        if tot and not sub and not iva:
            sub = round(tot / 1.16, 2)
            iva = round(tot - sub, 2)
            obs_parts.append("IVA desglosado")
        elif tot and iva and not sub:
            sub = round(tot - iva, 2)
        elif tot and sub and not iva:
            iva = round(tot - sub, 2)
            if iva < 0:
                iva = None
        elif sub and iva and not tot:
            tot = round(sub + iva, 2)
        elif calc_sub and sub and not iva and not tot:
            iva = round(sub * 0.16, 2)
            tot = round(sub + iva, 2)
    else:  # N/M
        if tot and iva and not sub:
            sub = round(tot - iva, 2)
        elif sub and iva and not tot:
            tot = round(sub + iva, 2)

    if pu is None and qty > 0:
        if sub is not None:
            pu = round(sub / qty, 2)
        elif tot is not None:
            if iva_f in ("Sí", "Incluido"):
                pu = round((tot / 1.16) / qty, 2)
            else:
                pu = round(tot / qty, 2)

    # Conversión Banxico
    tc: Optional[float] = None
    if moneda != "MXN" and bx_token:
        fecha_tc = fecha or datetime.date.today()
        tc = banxico_tc(moneda, fecha_tc, bx_token, on_warning=on_warning)
        if tc and tot is not None:
            tot_orig = tot
            tot = round(tot * tc, 2)
            sub = round(sub * tc, 2) if sub else None
            iva = round(iva * tc, 2) if iva else None
            pu  = round(pu  * tc, 2) if pu  else None
            obs_parts.append(f"1 {moneda} = ${tc:.4f} MXN")
            obs_parts.append(f"Total orig.: {moneda} {tot_orig:,.2f}")

    if tot is None and on_warning:
        on_warning(f"Sección '{label}': No se detectó un Total claro. Corrobora los montos manualmente.")

    return {
        "Nº Sec":                  sec_num,
        "Sección":                 label,
        "Tipo":                    tipo,
        "Fecha":                   fecha.isoformat() if fecha else datetime.date.today().isoformat(),
        "Rubro":                   label,
        "QT":                      "Sí",
        "T. Cambio":               moneda,
        "(+ IVA)":                 iva_f,
        "Cantidad":                qty,
        "Precio Unitario":         pu,
        "Subtotal (Sin IVA)":      sub,
        "IVA 16%":                 iva,
        "Total con IVA":           tot,
        "Diferencia final":        None,
        "Monto en Anexo Escrito":  None,
        "Observaciones":           " | ".join(obs_parts) if obs_parts else "",
    }


# ─────────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────────────────────
def _side(style: str = "thin") -> Side:
    return Side(style=style, color="000000")

def _border(style: str = "thin") -> Border:
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(rgb: str) -> PatternFill:
    return PatternFill("solid", start_color=rgb, end_color=rgb)


def to_excel(df: pd.DataFrame, nombre: str = "", blank: bool = False) -> bytes:
    """
    Genera Excel con formato profesional.
    blank=True produce plantilla vacía con fórmulas vivas.
    """
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = nombre[:31] if nombre else "Conciliación"

    F_WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    F_HDR   = Font(name="Calibri", size=11, bold=True, color="000000")
    F_DATA  = Font(name="Calibri", size=11, color="000000")
    F_TOT   = Font(name="Calibri", size=11, bold=True, color="000000")
    F_BOLD  = Font(name="Calibri", size=11, bold=True, color="000000")

    FILL_ROW1  = _fill("6E152E")
    FILL_HDR_A = _fill("D4C19C")
    FILL_HDR_B = _fill("EBE2D1")
    BD     = _border("thin")
    BD_MED = _border("medium")
    AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right",  vertical="center")

    # Configuraciones de impresión – Carta Horizontal
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.paperSize        = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation      = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage        = True
    ws.page_setup.fitToWidth       = 1   # encajar todas las columnas en 1 página de ancho
    ws.page_setup.fitToHeight      = 0   # sin límite de páginas en alto
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

    n_rows = len(df)

    for col_idx, width in enumerate(_WIDTHS_COL, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 27.75
    c = ws.cell(row=1, column=1, value=nombre or "")
    c.font = F_WHITE; c.fill = FILL_ROW1; c.border = BD; c.alignment = AL_L
    for ci in range(2, len(_COLS) + 1):
        cx = ws.cell(row=1, column=ci)
        cx.fill = FILL_ROW1; cx.border = BD

    ws.row_dimensions[2].height = 21.75
    for ci, hdr in enumerate(_COLS, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = F_HDR
        c.fill = FILL_HDR_A if ci <= 12 else FILL_HDR_B
        c.border = BD; c.alignment = AL_C

    DS = 3

    def _money_cell(ws_, r, ci, val=None):
        c_ = ws_.cell(row=r, column=ci, value=val)
        c_.number_format = _FMT_MONEY; c_.font = F_DATA
        c_.border = BD; c_.alignment = AL_R
        return c_

    for i, row_dict in enumerate(df.to_dict('records'), DS):
        r = i
        ws.row_dimensions[r].height = 16.5
        blank_row = blank

        c = ws.cell(row=r, column=1, value=str(row_dict.get("Tipo", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C
        c = ws.cell(row=r, column=2, value=str(row_dict.get("Fecha", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C
        c = ws.cell(row=r, column=3, value=str(row_dict.get("Rubro", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_L
        c = ws.cell(row=r, column=4, value=str(row_dict.get("QT", "Sí") or "Sí"))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C
        c = ws.cell(row=r, column=5, value=str(row_dict.get("T. Cambio", "MXN") or "MXN"))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C
        c = ws.cell(row=r, column=6, value=str(row_dict.get("(+ IVA)", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        q = _safe_f(row_dict.get("Cantidad"))
        c = ws.cell(row=r, column=7, value=int(q) if q is not None else 1)
        c.number_format = "0.00"; c.font = F_DATA; c.border = BD; c.alignment = AL_C

        pu_val = None if blank_row else _safe_f(row_dict.get("Precio Unitario"))
        _money_cell(ws, r, 8, pu_val)

        val_iva = str(row_dict.get("(+ IVA)", "N/M")).strip().lower().replace("\u00ed", "i").replace("\u00c3\u00ad", "i")
        has_iva = val_iva in ["si", "incluido"]

        c = ws.cell(row=r, column=9,  value=f'=IF(ISNUMBER(H{r}),G{r}*H{r},"")')
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        iva_formula = f'=IF(ISNUMBER(I{r}),I{r}*0.16,"")' if has_iva else None
        c = ws.cell(row=r, column=10, value=iva_formula)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        if has_iva:
            tot_formula = f'=IF(ISNUMBER(I{r}),IF(ISNUMBER(J{r}),I{r}+J{r},I{r}),"")'
        else:
            tot_formula = f'=IF(ISNUMBER(I{r}),I{r},"")'
        c = ws.cell(row=r, column=11, value=tot_formula)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        c = ws.cell(row=r, column=12, value=f'=IF(AND(ISNUMBER(K{r}),ISNUMBER(M{r})),K{r}-M{r},"")')
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        anx_val = None if blank_row else _safe_f(row_dict.get("Monto en Anexo Escrito"))
        c = _money_cell(ws, r, 13, anx_val); c.font = F_BOLD

        c = ws.cell(row=r, column=14, value="" if blank_row else str(row_dict.get("Observaciones", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_L

    if n_rows > 0:
        tot_row = DS + n_rows
        s_xl, e_xl = DS, DS + n_rows - 1
        ws.row_dimensions[tot_row].height = 18
        c = ws.cell(row=tot_row, column=1, value="TOTALES")
        c.font = F_TOT; c.border = BD_MED
        # Bordes en todas las celdas del rango de totales
        for ci in range(2, len(_COLS) + 1):
            cx = ws.cell(row=tot_row, column=ci)
            cx.border = BD_MED
        # SUMAs: Subtotal (9/I), IVA 16% (10/J), Total con IVA (11/K)
        for ci in (9, 10, 11):
            col_l = get_column_letter(ci)
            c = ws.cell(row=tot_row, column=ci, value=f"=SUM({col_l}{s_xl}:{col_l}{e_xl})")
            c.number_format = _FMT_MONEY; c.font = F_TOT
            c.fill = _fill("EBE2D1"); c.border = BD_MED; c.alignment = AL_R

    ws.freeze_panes = "A3"

    # Definir área de impresión explícita
    last_col = get_column_letter(len(_COLS))
    last_row = DS + n_rows if n_rows > 0 else DS
    ws.print_area = f"A1:{last_col}{last_row}"

    wb.save(buf)
    buf.seek(0)
    return buf.read()
